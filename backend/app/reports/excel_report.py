"""Excel report generator using openpyxl.

Produces an .xlsx workbook with 11 worksheets covering all result tabs.
"""

from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..api.schemas import ExportRequest
from . import _common as C

# ── Styles ───────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="1F4E79", size=11)
HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-size columns based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_kv_section(ws, title: str, rows: List, start_row: int) -> int:
    """Write a key-value section (label, value) and return next free row."""
    ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
    start_row += 1
    for i, (label, value) in enumerate(rows):
        if not label:  # spacer row
            start_row += 1
            continue
        r = start_row + i
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = Font(bold=True, name="Calibri", size=10)
        vc.alignment = Alignment(horizontal="right")
        lc.border = THIN_BORDER
        vc.border = THIN_BORDER
        if i % 2 == 0:
            lc.fill = ALT_FILL
            vc.fill = ALT_FILL
    return start_row + len(rows) + 1


def _write_table(ws, title: str, headers: List[str], rows: List[List[str]], start_row: int) -> int:
    """Write a table with headers and rows, return next free row."""
    ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
    start_row += 1
    # Headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    start_row += 1
    # Data rows
    for ri, row_data in enumerate(rows):
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            c.border = THIN_BORDER
            if ci > 1:
                c.alignment = Alignment(horizontal="right")
            if ri % 2 == 0:
                c.fill = ALT_FILL
    return start_row + len(rows) + 1


def generate_excel_report(buffer: BytesIO, req: ExportRequest) -> None:
    """Generate a complete Excel report into the provided buffer."""
    wb = Workbook()

    # ── 1. Summary ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Sulfidity Predictor V1 — Report").font = TITLE_FONT
    row = _write_kv_section(ws, "Key Inputs", C.get_key_inputs_rows(req), 3)
    row = _write_kv_section(ws, "Sulfidity", C.get_sulfidity_rows(req), row + 1)
    _auto_width(ws)

    # ── 2. WL Quality ────────────────────────────────────────────
    ws2 = wb.create_sheet("WL Quality")
    _write_kv_section(ws2, "White Liquor Quality", C.get_wl_quality_rows(req), 1)
    _auto_width(ws2)

    # ── 3. Makeup ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Makeup")
    _write_kv_section(ws3, "Makeup Chemical Requirements", C.get_makeup_rows(req), 1)
    _auto_width(ws3)

    # ── 4. Recovery Boiler ───────────────────────────────────────
    ws4 = wb.create_sheet("Recovery Boiler")
    _write_kv_section(ws4, "Recovery Boiler", C.get_recovery_boiler_rows(req), 1)
    _auto_width(ws4)

    # ── 5. Mass Balance ──────────────────────────────────────────
    ws5 = wb.create_sheet("Mass Balance")
    _write_kv_section(ws5, "Mass Balance", C.get_mass_balance_rows(req), 1)
    _auto_width(ws5)

    # ── 6. Inventory ─────────────────────────────────────────────
    ws6 = wb.create_sheet("Inventory")
    _write_kv_section(ws6, "Tank Inventory", C.get_inventory_rows(req), 1)
    _auto_width(ws6)

    # ── 7. Element Tracking ──────────────────────────────────────
    ws7 = wb.create_sheet("Element Tracking")
    _write_table(ws7, "Na/S Tracking by Unit Operation",
                 C.get_unit_operations_headers(), C.get_unit_operations_rows(req), 1)
    _auto_width(ws7)

    # ── 8. Loss Table ────────────────────────────────────────────
    ws8 = wb.create_sheet("Loss Table")
    _write_table(ws8, "Soda & Sulfur Losses",
                 C.get_loss_table_headers(), C.get_loss_table_rows(req), 1)
    _auto_width(ws8)

    # ── 9. Chemical Additions ────────────────────────────────────
    ws9 = wb.create_sheet("Chemical Additions")
    _write_table(ws9, "Chemical Additions",
                 C.get_chemical_additions_headers(), C.get_chemical_additions_rows(req), 1)
    _auto_width(ws9)

    # ── 10. Sensitivity (optional) ───────────────────────────────
    ws10 = wb.create_sheet("Sensitivity")
    sens_rows = C.get_sensitivity_rows(req.sensitivity_items)
    if sens_rows:
        _write_table(ws10, "Sensitivity Analysis",
                     C.get_sensitivity_headers(), sens_rows, 1)
    else:
        ws10.cell(row=1, column=1, value="No sensitivity data included").font = SUBTITLE_FONT
    _auto_width(ws10)

    # ── 11. Guidance ─────────────────────────────────────────────
    ws11 = wb.create_sheet("Guidance")
    guidance_rows = C.get_guidance_rows(req)
    if guidance_rows:
        _write_table(ws11, "Operational Guidance",
                     C.get_guidance_headers(), guidance_rows, 1)
    else:
        ws11.cell(row=1, column=1, value="No guidance items").font = SUBTITLE_FONT
    _auto_width(ws11)

    wb.save(buffer)
